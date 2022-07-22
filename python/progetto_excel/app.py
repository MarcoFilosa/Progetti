import openpyxl as xl #alias
from openpyxl.chart import BarChart, Reference


def process_workbook(filename):
    wb = xl.load_workbook(filename) #stiamo prendendo il file
    sheet = wb['Sheet1']

    #come funziona la raccolta dei box di excel
    cell = sheet['a1']
    sheet.cell(1, 1)
    #fine esempio

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        correct_price = 0.9 * cell.value  #valore corretto
        correct_price_cell = sheet.cell(row, 4) #posizione dove inserire il valore corretto
        correct_price_cell.value = correct_price #sccrivo nella posizione corretta il valore corretto

    #andriamo a considerare un gruppo di celle con Reference
    values = Reference(sheet,
              min_row=2,
              max_row=sheet.max_row,
              min_col=4,
              max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')

    wb.save(filename)

